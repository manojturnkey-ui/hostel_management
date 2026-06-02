from __future__ import annotations

from decimal import Decimal

from django.db import transaction

from openpyxl import load_workbook

from .models import (
    ActiveStatusChoices,
    Area,
    Building,
    Cot,
    CotStatusChoices,
    ExcelUploadLog,
    Floor,
    Room,
    Section,
)


EXPECTED_COLUMNS = [
    "area_name",
    "building_name",
    "section_name",
    "floor_name",
    "room_number",
    "cot_number",
    "cot_price",
    "security_deposit",
    "cot_status",
]


def normalize_cot_status(value: str | None) -> str:
    lookup = {
        "available": CotStatusChoices.AVAILABLE,
        "pending": CotStatusChoices.PENDING,
        "occupied": CotStatusChoices.OCCUPIED,
        "maintenance": CotStatusChoices.MAINTENANCE,
        "blocked": CotStatusChoices.BLOCKED,
    }
    return lookup.get((value or "").strip().lower(), CotStatusChoices.AVAILABLE)


def process_excel_upload(log_entry: ExcelUploadLog) -> ExcelUploadLog:
    uploaded_file = log_entry.uploaded_file
    uploaded_file.open("rb")
    workbook = load_workbook(uploaded_file, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    if headers != EXPECTED_COLUMNS:
        log_entry.status = "failed"
        log_entry.log_details = [f"Invalid columns. Expected: {', '.join(EXPECTED_COLUMNS)}"]
        log_entry.save(update_fields=["status", "log_details", "updated_at"])
        return log_entry

    success_count = 0
    failures: list[str] = []
    total_rows = 0

    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        total_rows += 1
        data = dict(zip(EXPECTED_COLUMNS, row))
        try:
            with transaction.atomic():
                area, _ = Area.objects.get_or_create(
                    area_name=str(data["area_name"]).strip(),
                    defaults={"status": ActiveStatusChoices.ACTIVE},
                )
                building, _ = Building.objects.get_or_create(
                    area=area,
                    building_name=str(data["building_name"]).strip(),
                    defaults={"status": ActiveStatusChoices.ACTIVE},
                )
                section, _ = Section.objects.get_or_create(
                    building=building,
                    section_name=str(data["section_name"]).strip(),
                    defaults={"status": ActiveStatusChoices.ACTIVE},
                )
                floor, _ = Floor.objects.get_or_create(
                    section=section,
                    floor_name=str(data["floor_name"]).strip(),
                    defaults={"status": ActiveStatusChoices.ACTIVE},
                )
                room, _ = Room.objects.get_or_create(
                    floor=floor,
                    room_number=str(data["room_number"]).strip(),
                    defaults={"status": ActiveStatusChoices.ACTIVE},
                )
                cot, created = Cot.objects.get_or_create(
                    room=room,
                    cot_number=str(data["cot_number"]).strip(),
                    defaults={
                        "cot_price": Decimal(str(data["cot_price"] or "0")),
                        "security_deposit": Decimal(str(data["security_deposit"] or "0")),
                        "status": normalize_cot_status(data["cot_status"]),
                    },
                )
                if not created:
                    cot.cot_price = Decimal(str(data["cot_price"] or "0"))
                    cot.security_deposit = Decimal(str(data["security_deposit"] or "0"))
                    cot.status = normalize_cot_status(data["cot_status"])
                    cot.save(update_fields=["cot_price", "security_deposit", "status", "updated_at"])
            success_count += 1
        except Exception as exc:
            failures.append(f"Row {row_index}: {exc}")

    log_entry.total_rows = total_rows
    log_entry.success_count = success_count
    log_entry.failure_count = len(failures)
    log_entry.status = "success" if not failures else ("partial" if success_count else "failed")
    log_entry.log_details = failures
    log_entry.save(update_fields=["total_rows", "success_count", "failure_count", "status", "log_details", "updated_at"])
    return log_entry
